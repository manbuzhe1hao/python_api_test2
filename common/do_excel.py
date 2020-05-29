from openpyxl import load_workbook

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
    def get_data(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        case_data=[]
        for i in range(2,sheet.max_row+1):
            case = {}
            case['case_id']=sheet.cell(row=i,column=1).value
            case['title']=sheet.cell(row=i,column=2).value
            case['url']=sheet.cell(row=i,column=3).value
            case['data']=sheet.cell(row=i,column=4).value
            case['method']=sheet.cell(row=i,column=5).value
            case['expected']=sheet.cell(row=i,column=6).value
            case_data.append(case)
        return case_data
    def write_data(self,row,col,value):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row,col).value=value
        wb.save(self.file_name)

if __name__ == '__main__':
    from common import concant
    do_excel=DoExcel(concant.excel_dir,'register').get_data()
    print(do_excel)




